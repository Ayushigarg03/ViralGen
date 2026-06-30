import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, case
from .models import Campaign, Job
from .database import SessionLocal

def generate_campaign_excel(user_id: int) -> BytesIO:
    """
    Generates a beautifully formatted multi-sheet Excel file containing:
    1. Campaign History
    2. Analytics Summary
    3. Daily Trends
    Returns a BytesIO binary buffer of the generated file.
    """
    db = SessionLocal()
    wb = Workbook()
    
    # Define styles
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo brand color
    
    success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Mint green
    failed_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Pastel red
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")   # Light gray
    
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # -------------------------------------------------------------
    # SHEET 1: Campaign History
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Campaign History"
    
    headers1 = [
        "Campaign ID", "Date Created", "Platform", 
        "Brand Persona", "Brand Brief", "Generated Copy", 
        "Image URL", "Status"
    ]
    ws1.append(headers1)
    
    # Query campaigns
    campaigns = db.query(Campaign).filter(Campaign.user_id == user_id).order_by(Campaign.created_at.desc()).all()
    
    for c in campaigns:
        ws1.append([
            c.id,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
            c.platform,
            c.persona,
            c.brief,
            c.text_copy or "",
            c.image_url or "",
            c.status
        ])
        
    # Style Sheet 1
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border
        
    for row_idx in range(2, len(campaigns) + 2):
        status_val = ws1.cell(row=row_idx, column=8).value
        # Determine row fill
        row_fill = None
        if status_val == "SUCCESS":
            row_fill = success_fill
        elif status_val == "FAILED":
            row_fill = failed_fill
        elif row_idx % 2 == 0:
            row_fill = zebra_fill
            
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            # Alignments
            if col_idx in [1, 2, 3, 4, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # -------------------------------------------------------------
    # SHEET 2: Analytics Summary
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Analytics Summary")
    headers2 = ["Platform", "Total Campaigns", "Success Rate (%)"]
    ws2.append(headers2)
    
    # Query analytics data grouped by platform
    platform_data = db.query(
        Campaign.platform,
        func.count(Campaign.id).label("total"),
        func.sum(case((Campaign.status == 'SUCCESS', 1), else_=0)).label("success")
    ).filter(Campaign.user_id == user_id).group_by(Campaign.platform).all()
    
    for row in platform_data:
        platform, total, success = row
        rate = round((success / total * 100), 2) if total > 0 else 0.0
        ws2.append([platform, total, rate])
        
    # Style Sheet 2
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx in range(2, len(platform_data) + 2):
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    # -------------------------------------------------------------
    # SHEET 3: Daily Trends
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Daily Trends")
    headers3 = ["Date", "Total Generated", "Successful", "Failed"]
    ws3.append(headers3)
    
    # Query daily stats
    # Truncate to date
    daily_stats = db.query(
        func.date(Campaign.created_at).label("date_val"),
        func.count(Campaign.id).label("total"),
        func.sum(case((Campaign.status == 'SUCCESS', 1), else_=0)).label("success"),
        func.sum(case((Campaign.status == 'FAILED', 1), else_=0)).label("failed")
    ).filter(Campaign.user_id == user_id).group_by("date_val").order_by("date_val").all()
    
    for stat in daily_stats:
        date_str = str(stat[0]) if stat[0] else ""
        ws3.append([
            date_str,
            stat[1],
            stat[2] or 0,
            stat[3] or 0
        ])
        
    # Style Sheet 3
    for col_idx in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx in range(2, len(daily_stats) + 2):
        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    # -------------------------------------------------------------
    # Auto-adjust column widths across all sheets
    # -------------------------------------------------------------
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Check for wrapped text length or cell string value length
                if cell.value:
                    val_str = str(cell.value)
                    # Limit the calculation width for extremely long generated copies
                    if len(val_str) > 50:
                        max_len = max(max_len, 50)
                    else:
                        max_len = max(max_len, len(val_str))
            # Set width with padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Specially format rows for headers and sheets
    ws1.row_dimensions[1].height = 28
    ws2.row_dimensions[1].height = 28
    ws3.row_dimensions[1].height = 28
    
    db.close()
    
    # Save to buffer
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
